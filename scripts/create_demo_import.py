#!/usr/bin/env python
"""Create a small XLSX import file that intentionally triggers the demo pipeline."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_DIR = Path(__file__).resolve().parents[1]
SOURCE_CSV = (
    REPO_DIR
    / "data/raw/databricks_virtue_foundation_dataset_dais_2026"
    / "virtue_foundation_dataset/facilities/facilities.csv.gz"
)
OUTPUT_DIR = REPO_DIR / "demo"
OUTPUT_XLSX = OUTPUT_DIR / "data_readiness_demo_import.xlsx"


COLUMNS = [
    "unique_id",
    "name",
    "organization_type",
    "officialPhone",
    "email",
    "websites",
    "yearEstablished",
    "address_line1",
    "address_city",
    "address_stateOrRegion",
    "address_zipOrPostcode",
    "address_country",
    "specialties",
    "procedure",
    "equipment",
    "capability",
    "description",
    "numberDoctors",
    "capacity",
    "latitude",
    "longitude",
    "cluster_id",
    "source",
    "source_urls",
    "demo_trigger",
    "expected_agent_signal",
]


def source_row(df: pd.DataFrame, name: str) -> dict:
    match = df[df["name"].fillna("").astype(str).eq(name)]
    if match.empty:
        raise ValueError(f"Could not find source row: {name}")
    row = match.iloc[0].fillna("").to_dict()
    return {column: row.get(column, "") for column in COLUMNS}


def row(base: dict, **updates) -> dict:
    out = {column: base.get(column, "") for column in COLUMNS}
    out.update(updates)
    return out


def build_rows() -> list[dict]:
    df = pd.read_csv(SOURCE_CSV, dtype=str).fillna("")

    fortis = source_row(df, "Fortis Hospital, Gurugram")
    aiims = source_row(df, "All India Institute of Medical Sciences Patna")
    hcg = source_row(df, "HCG Manavata Cancer Centre")
    civil = source_row(df, "Civil Hospital")

    return [
        row(
            fortis,
            unique_id="demo-import-001",
            capability='["Emergency", "ICU", "trauma"]',
            description="Imported exact-name update claiming emergency, trauma, and ICU services.",
            demo_trigger="Exact existing-name duplicate",
            expected_agent_signal="DedupAgent duplicate decision; HumanReviewGate proof/reject item",
        ),
        row(
            fortis,
            unique_id="demo-import-002",
            name="Fortis Hospital Gurgaon",
            address_zipOrPostcode="122002",
            capability='["24x7 emergency", "ventilator support"]',
            description="Near-duplicate name variation for the same Gurgaon facility.",
            demo_trigger="Near duplicate with changed name",
            expected_agent_signal="LLM dedupe review when enabled; deterministic mode inserts but demo notes call it out",
        ),
        row(
            aiims,
            unique_id="demo-import-003",
            capability='["Maternity", "NICU", "Emergency"]',
            description="Exact AIIMS Patna row with added NICU and emergency claims.",
            demo_trigger="Exact existing-name duplicate with new capability claims",
            expected_agent_signal="DedupAgent duplicate decision; evidence/provenance question for reviewer",
        ),
        row(
            aiims,
            unique_id="demo-import-004",
            name="AIIMS Patna Emergency Annex",
            address_zipOrPostcode="801105",
            capability='["Emergency surgery", "trauma stabilization"]',
            description="Annex record appears related to AIIMS Patna but may be a separate site.",
            demo_trigger="Ambiguous related facility",
            expected_agent_signal="Ingest insert/review candidate; planning impact discussion",
        ),
        row(
            {},
            unique_id="demo-import-005",
            name="North District Maternity Centre",
            organization_type="facility",
            address_line1="Near bus stand, ward 12",
            address_city="Patna",
            address_stateOrRegion="Bihar",
            address_zipOrPostcode="",
            address_country="India",
            specialties='["obstetrics", "gynecology"]',
            capability='["Maternity", "NICU claimed"]',
            description="Delivery services listed. NICU mentioned once, no equipment evidence supplied.",
            latitude="",
            longitude="",
            cluster_id="demo-import-maternity",
            source="demo_xls",
            source_urls='["https://example.org/demo/north-maternity"]',
            demo_trigger="Sparse location plus weak NICU claim",
            expected_agent_signal="Ingestion row quality flag; review item for missing PIN/geocode",
        ),
        row(
            {},
            unique_id="demo-import-006",
            name="Shree Maternity and NICU",
            organization_type="facility",
            address_line1="Old market road",
            address_city="",
            address_stateOrRegion="Rajasthan",
            address_zipOrPostcode="302004",
            address_country="India",
            specialties='["obstetrics"]',
            capability='["NICU", "newborn care"]',
            description="NICU claim appears in capability field, but city is blank.",
            latitude="26.91",
            longitude="75.79",
            cluster_id="demo-import-maternity",
            source="demo_xls",
            source_urls='["https://example.org/demo/shree-nicu"]',
            demo_trigger="Missing city on clinically important NICU row",
            expected_agent_signal="Ingestion row quality flag; HumanReviewGate item",
        ),
        row(
            {},
            unique_id="demo-import-007",
            name="City Care Hospital",
            organization_type="facility",
            address_line1="MI Road",
            address_city="Jaipur",
            address_stateOrRegion="Rajasthan",
            address_zipOrPostcode="302001",
            address_country="India",
            specialties='["emergencyMedicine", "internalMedicine"]',
            capability='["24x7 Emergency Services"]',
            description="Multispecialty hospital with emergency department.",
            latitude="26.9124",
            longitude="75.7873",
            cluster_id="demo-import-citycare",
            source="demo_xls",
            source_urls='["https://example.org/demo/city-care"]',
            demo_trigger="Incoming duplicate pair row 1",
            expected_agent_signal="Duplicate cluster story for demo; LLM mode should review/merge",
        ),
        row(
            {},
            unique_id="demo-import-008",
            name="City Care Hosp.",
            organization_type="facility",
            address_line1="M I Road",
            address_city="Jaipur",
            address_stateOrRegion="Rajasthan",
            address_zipOrPostcode="302001",
            address_country="India",
            specialties='["emergencyMedicine"]',
            capability='["Emergency", "trauma"]',
            description="Emergency and trauma services; likely same as City Care Hospital.",
            latitude="26.9125",
            longitude="75.7874",
            cluster_id="demo-import-citycare",
            source="demo_xls",
            source_urls='["https://example.org/demo/city-care-alt"]',
            demo_trigger="Incoming duplicate pair row 2",
            expected_agent_signal="Duplicate cluster story for demo; LLM mode should review/merge",
        ),
        row(
            hcg,
            unique_id="demo-import-009",
            capability='["Oncology", "chemotherapy", "radiation oncology"]',
            description="Exact existing oncology centre row with stronger structured oncology claims.",
            demo_trigger="Exact existing-name duplicate with stronger evidence",
            expected_agent_signal="DedupAgent duplicate decision; possible update candidate when LLM enabled",
        ),
        row(
            hcg,
            unique_id="demo-import-010",
            name="HCG Manavata Cancer Centre Nashik",
            address_zipOrPostcode="422002",
            capability='["Oncology", "chemotherapy"]',
            description="Same oncology centre with city appended to the name.",
            demo_trigger="Near duplicate oncology row",
            expected_agent_signal="LLM dedupe review when enabled; deterministic mode inserts",
        ),
        row(
            civil,
            unique_id="demo-import-011",
            yearEstablished="1750",
            address_zipOrPostcode="",
            description="Civil hospital import with impossible/stale year and missing postcode.",
            demo_trigger="Suspicious metadata plus missing PIN",
            expected_agent_signal="Ingestion row quality flag; QA metadata story for later incoming QA",
        ),
        row(
            {},
            unique_id="demo-import-012",
            name="Rural Trauma Stabilization Unit",
            organization_type="facility",
            address_line1="Highway checkpoint",
            address_city="Barmer",
            address_stateOrRegion="Rajasthan",
            address_zipOrPostcode="344001",
            address_country="India",
            specialties="",
            capability='["Emergency", "trauma", "surgery"]',
            description="Claimed trauma stabilization and surgery, but no doctor count, capacity, or equipment detail.",
            numberDoctors="",
            capacity="",
            latitude="",
            longitude="",
            cluster_id="demo-import-trauma",
            source="demo_xls",
            source_urls='["https://example.org/demo/rural-trauma"]',
            demo_trigger="High-risk claim with sparse operational fields",
            expected_agent_signal="Risk/planning story: count carefully until evidence improves",
        ),
    ]


def write_workbook(rows: list[dict]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    notes = [
        {
            "demo_step": 1,
            "what_to_show": "Upload this workbook in Import + Actions.",
            "expected_result": "Preview parses 12 rows and required source columns.",
        },
        {
            "demo_step": 2,
            "what_to_show": "Run ingestion pipeline.",
            "expected_result": "All ten agents complete in local skeleton mode, including PIN and NFHS context checks.",
        },
        {
            "demo_step": 3,
            "what_to_show": "Open pipeline agent cards.",
            "expected_result": "Dedup sees exact existing-name duplicates; review sees missing required row values.",
        },
        {
            "demo_step": 4,
            "what_to_show": "Tell the Track 4 -> Track 2 story.",
            "expected_result": "The import creates proof/reject work before trusted data feeds risk planning.",
        },
    ]

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=COLUMNS).to_excel(writer, sheet_name="Facilities_Import", index=False)
        pd.DataFrame(notes).to_excel(writer, sheet_name="Demo_Notes", index=False)

    wb = load_workbook(OUTPUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    trigger_fill = PatternFill("solid", fgColor="FEF3C7")
    required_fill = PatternFill("solid", fgColor="DBEAFE")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = str(column_cells[0].value or "")
            max_len = max(len(str(cell.value or "")) for cell in column_cells[: min(len(column_cells), 20)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 44)
            if header in {"name", "address_city", "address_stateOrRegion", "address_zipOrPostcode"}:
                column_cells[0].fill = required_fill
            if header in {"demo_trigger", "expected_agent_signal"}:
                column_cells[0].fill = trigger_fill

    wb.save(OUTPUT_XLSX)


def main() -> None:
    rows = build_rows()
    write_workbook(rows)
    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
